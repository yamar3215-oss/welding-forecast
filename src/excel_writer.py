"""Excelレポート生成: 6シート + 補助シート構成."""
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, BarChart, Reference


_EVAL_FILL = {
    1: PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    2: PatternFill(start_color="FFD8A8", end_color="FFD8A8", fill_type="solid"),
    3: PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    4: PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
    5: PatternFill(start_color="D9C6E5", end_color="D9C6E5", fill_type="solid"),
}


def write_report(
    out_path: str,
    unit_consumption: pd.DataFrame,
    forecast_monthly: pd.DataFrame,
    per_ship_forecast: pd.DataFrame,
    order_recommend: pd.DataFrame,
    stock_eval: pd.DataFrame,
    unmapped_skus: pd.DataFrame,
    warnings_list: list,
    mape: float = float("nan"),
    consumption_history: pd.DataFrame = None,
    eval_compare: dict = None,
    mape_baseline: float = float("nan"),
    mape_per_sku: pd.DataFrame = None,
    mape_per_sku_baseline: pd.DataFrame = None,
    order_validation: dict = None,
    materials: pd.DataFrame = None,
    shape: dict | None = None,
) -> None:
    """6+補助シートのExcelレポートを生成."""
    if consumption_history is None:
        consumption_history = pd.DataFrame(columns=["year_month", "sku_id", "qty_kg"])
    if eval_compare is None:
        eval_compare = {"total": 0, "matched": 0, "match_rate": float("nan"),
                        "mismatches": pd.DataFrame()}
    if mape_per_sku is None:
        mape_per_sku = pd.DataFrame(columns=["sku_id", "mape_%"])
    if mape_per_sku_baseline is None:
        mape_per_sku_baseline = pd.DataFrame(columns=["sku_id", "mape_%"])
    wb = Workbook()
    wb.remove(wb.active)

    _write_overview(
        wb, forecast_monthly, warnings_list,
        mape=mape, consumption_history=consumption_history,
        eval_compare=eval_compare,
        mape_baseline=mape_baseline,
        order_validation=order_validation,
    )
    _write_unit_consumption(wb, unit_consumption)
    _write_monthly_forecast(wb, forecast_monthly)
    _write_per_ship_forecast(wb, per_ship_forecast)
    _write_order_recommend(wb, order_recommend)
    _write_stock_eval(wb, stock_eval)
    _write_evaluation_mismatch(wb, eval_compare)
    _write_mape_comparison(wb, mape_per_sku, mape_per_sku_baseline)
    if order_validation is not None:
        _write_order_validation(wb, order_validation, materials)
    if shape is not None:
        _write_category_shape(wb, shape or {}, materials)
    _write_unmapped(wb, unmapped_skus)
    _write_warnings(wb, warnings_list)

    wb.save(out_path)


def _write_overview(wb, forecast_monthly, warnings_list, mape=float("nan"),
                    consumption_history=None, eval_compare=None,
                    mape_baseline=float("nan"), order_validation=None):
    if consumption_history is None:
        consumption_history = pd.DataFrame(columns=["year_month", "sku_id", "qty_kg"])
    if eval_compare is None:
        eval_compare = {"total": 0, "matched": 0, "match_rate": float("nan"), "mismatches": pd.DataFrame()}
    ws = wb.create_sheet("概要")
    ws["A1"] = "船舶溶接材料 予測レポート"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A3"] = "生成日時"
    ws["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws["A4"] = "予測期間"
    if len(forecast_monthly) > 0:
        ws["B4"] = f"{forecast_monthly['year_month'].min()} 〜 {forecast_monthly['year_month'].max()}"
    ws["A5"] = "予測総使用量 [kg]"
    ws["B5"] = float(forecast_monthly["予測使用量"].sum()) if len(forecast_monthly) > 0 else 0
    ws["A6"] = "警告件数"
    ws["B6"] = len(warnings_list)
    ws["A7"] = "ホールドアウト MAPE [%] (制約あり)"
    ws["B7"] = mape if not pd.isna(mape) else "N/A"
    ws["A8"] = "ホールドアウト MAPE [%] (制約なし)"
    ws["B8"] = mape_baseline if not pd.isna(mape_baseline) else "N/A"
    ws["A9"] = "MAPE改善幅 [pt]"
    if not pd.isna(mape) and not pd.isna(mape_baseline):
        ws["B9"] = mape_baseline - mape
    else:
        ws["B9"] = "N/A"
    ws["A10"] = "既存在庫評価との一致率 [%]"
    if eval_compare["total"] > 0:
        ws["B10"] = eval_compare["match_rate"] * 100
        ws["C10"] = f"{eval_compare['matched']}/{eval_compare['total']}"
    else:
        ws["B10"] = "N/A"

    # 発注検証 (_2.xlsx) ブロック
    if order_validation is not None:
        ws["A11"] = "発注検証 MAPE [%] (推奨 vs 入荷予定)"
        ws["B11"] = order_validation["mape_overall"] if not pd.isna(order_validation["mape_overall"]) else "N/A"
        cov = order_validation["coverage"]
        ws["A12"] = "発注検証 SKUカバー率"
        ws["B12"] = f"{cov['captured']}/{cov['planned_skus']}"
        ws["C12"] = f"FP:{cov['false_positive_skus']} / 漏れ:{cov['missed_skus']}"

    # 月別総消費推移チャート (行番号は13/14 に下がる)
    history = consumption_history.groupby("year_month")["qty_kg"].sum().reset_index() if len(consumption_history) > 0 else pd.DataFrame(columns=["year_month", "qty_kg"])
    history = history.sort_values("year_month").tail(24)
    history = history.rename(columns={"qty_kg": "実績"})
    forecast_total = forecast_monthly.groupby("year_month")["予測使用量"].sum().reset_index() if len(forecast_monthly) > 0 else pd.DataFrame(columns=["year_month", "予測使用量"])
    forecast_total = forecast_total.rename(columns={"予測使用量": "予測"})
    combined = pd.merge(history, forecast_total, on="year_month", how="outer").sort_values("year_month")
    ws["A13"] = "年月"
    ws["B13"] = "実績"
    ws["C13"] = "予測"
    for i, row in enumerate(combined.itertuples(index=False), start=14):
        ws.cell(row=i, column=1, value=row.year_month)
        ws.cell(row=i, column=2, value=float(row.実績) if pd.notna(row.実績) else None)
        ws.cell(row=i, column=3, value=float(row.予測) if pd.notna(row.予測) else None)
    end_row = 13 + len(combined)
    if end_row > 13:
        chart = LineChart()
        chart.title = "月別総消費トレンド (実績＋予測)"
        chart.y_axis.title = "kg"
        chart.x_axis.title = "年月"
        data = Reference(ws, min_col=2, min_row=13, max_col=3, max_row=end_row)
        cats = Reference(ws, min_col=1, min_row=14, max_row=end_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 10
        chart.width = 20
        ws.add_chart(chart, "E3")


def _write_unit_consumption(wb, unit):
    ws = wb.create_sheet("船種別原単位")
    if len(unit) == 0:
        ws["A1"] = "(データなし)"
        return
    pivot = unit.pivot_table(
        index="ship_type", columns="sku_id", values="kg_per_ship", aggfunc="first"
    ).fillna(0.0)
    _write_dataframe_with_index(ws, pivot, header_label="船種＼SKU")


def _write_monthly_forecast(wb, forecast):
    ws = wb.create_sheet("月次予測使用量")
    if len(forecast) == 0:
        ws["A1"] = "(データなし)"
        return
    top10 = (
        forecast.groupby("sku_id")["予測使用量"].sum()
        .sort_values(ascending=False).head(10).index.tolist()
    )
    fc_top = forecast[forecast["sku_id"].isin(top10)]
    pivot = fc_top.pivot_table(
        index="year_month", columns="sku_id", values="予測使用量", aggfunc="sum"
    ).fillna(0.0)
    _write_dataframe_with_index(ws, pivot, header_label="年月＼SKU(上位10)")
    n_rows, n_cols = pivot.shape
    if n_rows > 0 and n_cols > 0:
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "stacked"
        chart.overlap = 100
        chart.title = "月次予測使用量 (上位10SKU積み上げ)"
        chart.y_axis.title = "kg"
        data = Reference(ws, min_col=2, min_row=1, max_col=n_cols + 1, max_row=n_rows + 1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=n_rows + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 10
        chart.width = 18
        ws.add_chart(chart, f"A{n_rows + 4}")


def _write_per_ship_forecast(wb, per_ship):
    ws = wb.create_sheet("船別予測使用量")
    if len(per_ship) == 0:
        ws["A1"] = "(データなし)"
        return
    pivot = per_ship.pivot_table(
        index=["s_no", "ship_type"], columns="sku_id", values="予測使用量", aggfunc="sum"
    ).fillna(0.0)
    _write_dataframe_with_index(ws, pivot, header_label="船番/船種＼SKU")


def _write_order_recommend(wb, order):
    ws = wb.create_sheet("推奨発注量")
    if len(order) == 0:
        ws["A1"] = "(データなし)"
        return
    pivot = order.pivot_table(
        index="sku_id", columns="year_month", values="推奨発注量", aggfunc="sum"
    ).fillna(0.0)
    _write_dataframe_with_index(ws, pivot, header_label="SKU＼年月")

    # 詳細セクション: 月末予測在庫と予測在庫評価 (SKUごとに横並び)
    detail_start = len(pivot) + 4
    ws.cell(row=detail_start, column=1, value="月末予測在庫トラジェクトリ").font = Font(bold=True)
    stock_pivot = order.pivot_table(
        index="sku_id", columns="year_month", values="月末予測在庫", aggfunc="sum"
    ).fillna(0.0)
    _write_dataframe_with_index_at(ws, stock_pivot, header_label="SKU＼年月",
                                    start_row=detail_start + 1)

    eval_start = detail_start + 2 + len(stock_pivot)
    ws.cell(row=eval_start, column=1, value="予測在庫評価 (1=過少 / 3=適正 / 5=過剰)").font = Font(bold=True)
    eval_pivot = order.pivot_table(
        index="sku_id", columns="year_month", values="予測在庫評価", aggfunc="first"
    )
    _write_dataframe_with_index_at(ws, eval_pivot, header_label="SKU＼年月",
                                    start_row=eval_start + 1)
    # 評価値の色塗り
    n_rows, n_cols = eval_pivot.shape
    for r in range(n_rows):
        for c in range(n_cols):
            cell = ws.cell(row=eval_start + 2 + r, column=c + 2)
            val = eval_pivot.iat[r, c]
            if pd.notna(val) and int(val) in _EVAL_FILL:
                cell.fill = _EVAL_FILL[int(val)]

    # 月別総発注量チャート
    totals_by_month = order.groupby("year_month")["推奨発注量"].sum().reset_index()
    summary_start_row = eval_start + 4 + len(eval_pivot)
    ws.cell(row=summary_start_row, column=1, value="月別総発注量").font = Font(bold=True)
    ws.cell(row=summary_start_row + 1, column=1, value="年月")
    ws.cell(row=summary_start_row + 1, column=2, value="推奨発注量合計")
    for i, row in enumerate(totals_by_month.itertuples(index=False), start=summary_start_row + 2):
        ws.cell(row=i, column=1, value=row.year_month)
        ws.cell(row=i, column=2, value=float(row.推奨発注量))
    end_row = summary_start_row + 1 + len(totals_by_month)
    if len(totals_by_month) > 0:
        chart = BarChart()
        chart.type = "col"
        chart.title = "月別総発注量"
        chart.y_axis.title = "kg"
        data = Reference(ws, min_col=2, min_row=summary_start_row + 1, max_row=end_row)
        cats = Reference(ws, min_col=1, min_row=summary_start_row + 2, max_row=end_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 10
        chart.width = 18
        ws.add_chart(chart, f"D{summary_start_row}")


def _write_stock_eval(wb, stock_eval):
    ws = wb.create_sheet("在庫評価")
    if len(stock_eval) == 0:
        ws["A1"] = "(データなし)"
        return
    pivot = stock_eval.pivot_table(
        index="sku_id", columns="year_month", values="評価", aggfunc="first"
    )
    _write_dataframe_with_index(ws, pivot, header_label="SKU＼年月")
    n_rows, n_cols = pivot.shape
    for r in range(n_rows):
        for c in range(n_cols):
            cell = ws.cell(row=r + 2, column=c + 2)
            val = pivot.iat[r, c]
            if pd.notna(val) and int(val) in _EVAL_FILL:
                cell.fill = _EVAL_FILL[int(val)]


def _write_evaluation_mismatch(wb, eval_compare):
    """既存在庫評価と計算値が一致しなかったレコードを表示."""
    ws = wb.create_sheet("評価不一致")
    mismatches = eval_compare["mismatches"]
    if len(mismatches) == 0:
        ws["A1"] = "全件一致 (または比較対象なし)"
        return
    ws["A1"] = "SKU"
    ws["B1"] = "年月"
    ws["C1"] = "計算値"
    ws["D1"] = "既存値"
    for col in "ABCD":
        ws[f"{col}1"].font = Font(bold=True)
    for i, row in enumerate(mismatches.itertuples(index=False), start=2):
        ws.cell(row=i, column=1, value=row.sku_id)
        ws.cell(row=i, column=2, value=row.year_month)
        ws.cell(row=i, column=3, value=int(row.computed))
        ws.cell(row=i, column=4, value=float(row.existing))


def _write_mape_comparison(wb, mape_per_sku, mape_per_sku_baseline):
    """SKU別MAPEの「制約あり/なし/改善」比較シート."""
    ws = wb.create_sheet("MAPE比較")
    if len(mape_per_sku) == 0 and len(mape_per_sku_baseline) == 0:
        ws["A1"] = "(MAPE算出不可)"
        return

    base = mape_per_sku_baseline.rename(columns={"mape_%": "MAPE_制約なし"})
    with_c = mape_per_sku.rename(columns={"mape_%": "MAPE_制約あり"})
    merged = pd.merge(
        base[["sku_id", "MAPE_制約なし"]],
        with_c[["sku_id", "MAPE_制約あり"]],
        on="sku_id", how="outer",
    )
    merged["改善_pt"] = merged["MAPE_制約なし"] - merged["MAPE_制約あり"]
    merged["銘柄"] = merged["sku_id"].apply(lambda s: str(s).split("_", 1)[0])
    merged = merged.sort_values("改善_pt", ascending=False, na_position="last")

    headers = ["sku_id", "銘柄", "MAPE_制約なし [%]", "MAPE_制約あり [%]", "改善 [pt]"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h).font = Font(bold=True)
    for i, row in enumerate(merged.itertuples(index=False), start=2):
        ws.cell(row=i, column=1, value=str(row.sku_id))
        ws.cell(row=i, column=2, value=str(row.銘柄))
        ws.cell(row=i, column=3,
                value=float(row.MAPE_制約なし) if pd.notna(row.MAPE_制約なし) else None)
        ws.cell(row=i, column=4,
                value=float(row.MAPE_制約あり) if pd.notna(row.MAPE_制約あり) else None)
        ws.cell(row=i, column=5,
                value=float(row.改善_pt) if pd.notna(row.改善_pt) else None)


def _write_order_validation(wb, result: dict, materials: pd.DataFrame = None):
    """推奨発注 vs 入荷予定 の検証結果シート."""
    ws = wb.create_sheet("発注検証_2026Q2")
    cmp = result["comparison"]
    if len(cmp) == 0:
        ws["A1"] = "(検証データなし)"
        return

    months = sorted(cmp["year_month"].unique())
    # 銘柄ルックアップ
    brand_map = {}
    if materials is not None and len(materials) > 0:
        brand_map = dict(zip(materials["sku_id"], materials["銘柄"]))

    # ヘッダ: sku_id, 銘柄, 月ごとに (推奨, 計画, 誤差%), SKU別MAPE
    headers = ["sku_id", "銘柄"]
    for m in months:
        headers += [f"{m} 推奨", f"{m} 計画", f"{m} 誤差%"]
    headers += ["SKU別MAPE [%]", "n_months"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h).font = Font(bold=True)

    per_sku = result["mape_per_sku"].set_index("sku_id")
    skus = sorted(cmp["sku_id"].unique())
    # 表示順: planned が大きい SKU を上に
    plan_total = cmp.groupby("sku_id")["planned_kg"].sum().sort_values(ascending=False)
    skus = plan_total.index.tolist() + [s for s in skus if s not in plan_total.index]

    row_idx = 2
    for sku in skus:
        ws.cell(row=row_idx, column=1, value=sku)
        ws.cell(row=row_idx, column=2, value=brand_map.get(sku, ""))
        col_idx = 3
        for m in months:
            sub = cmp[(cmp["sku_id"] == sku) & (cmp["year_month"] == m)]
            if len(sub) > 0:
                r = sub.iloc[0]
                ws.cell(row=row_idx, column=col_idx, value=float(r["recommended_kg"]))
                ws.cell(row=row_idx, column=col_idx + 1, value=float(r["planned_kg"]))
                ape = r["abs_pct_error"]
                if pd.notna(ape):
                    ws.cell(row=row_idx, column=col_idx + 2, value=float(ape))
            col_idx += 3
        if sku in per_sku.index:
            ws.cell(row=row_idx, column=col_idx, value=float(per_sku.loc[sku, "mape"]))
            ws.cell(row=row_idx, column=col_idx + 1, value=int(per_sku.loc[sku, "n_months"]))
        row_idx += 1

    # 集計ブロック (2行空けて)
    summary_row = row_idx + 2
    ws.cell(row=summary_row, column=1, value="=== 集計 ===").font = Font(bold=True)
    summary_row += 1

    ws.cell(row=summary_row, column=1, value="aggregate MAPE [%]")
    ws.cell(row=summary_row, column=2, value=float(result["mape_overall"]) if not pd.isna(result["mape_overall"]) else "N/A")
    summary_row += 1
    ws.cell(row=summary_row, column=1, value="±20% 一致率 [%]")
    ws.cell(row=summary_row, column=2, value=float(result["within_20pct_rate"]) * 100 if not pd.isna(result["within_20pct_rate"]) else "N/A")
    summary_row += 1
    cov = result["coverage"]
    ws.cell(row=summary_row, column=1, value="SKUカバー率 (captured / planned)")
    ws.cell(row=summary_row, column=2, value=f"{cov['captured']}/{cov['planned_skus']}")
    ws.cell(row=summary_row, column=3, value=f"カバー率 {cov['captured_rate']*100:.1f}%" if not pd.isna(cov['captured_rate']) else "N/A")
    summary_row += 1
    ws.cell(row=summary_row, column=1, value="false positive SKU (推奨あるが計画なし)")
    ws.cell(row=summary_row, column=2, value=int(cov["false_positive_skus"]))
    summary_row += 1
    ws.cell(row=summary_row, column=1, value="missed SKU (計画あるが推奨なし)")
    ws.cell(row=summary_row, column=2, value=int(cov["missed_skus"]))
    summary_row += 1

    tot = result["totals"]
    summary_row += 1
    ws.cell(row=summary_row, column=1, value="推奨総量 [kg]")
    ws.cell(row=summary_row, column=2, value=float(tot["recommended_total_kg"]))
    summary_row += 1
    ws.cell(row=summary_row, column=1, value="計画総量 [kg]")
    ws.cell(row=summary_row, column=2, value=float(tot["planned_total_kg"]))
    summary_row += 1
    ws.cell(row=summary_row, column=1, value="比 (推奨/計画)")
    ws.cell(row=summary_row, column=2, value=float(tot["ratio"]) if not pd.isna(tot["ratio"]) else "N/A")
    summary_row += 1

    summary_row += 1
    ws.cell(row=summary_row, column=1, value="=== 月別MAPE ===").font = Font(bold=True)
    summary_row += 1
    ws.cell(row=summary_row, column=1, value="年月").font = Font(bold=True)
    ws.cell(row=summary_row, column=2, value="MAPE [%]").font = Font(bold=True)
    ws.cell(row=summary_row, column=3, value="比較SKU数").font = Font(bold=True)
    for _, mrow in result["mape_per_month"].iterrows():
        summary_row += 1
        ws.cell(row=summary_row, column=1, value=mrow["year_month"])
        ws.cell(row=summary_row, column=2, value=float(mrow["mape"]) if pd.notna(mrow["mape"]) else None)
        ws.cell(row=summary_row, column=3, value=int(mrow["n_skus"]))


def _write_category_shape(wb, shape: dict, materials: pd.DataFrame | None):
    """SKU区分別 shape factor の表示シート."""
    ws = wb.create_sheet("shape_factor")
    if not shape:
        ws["A1"] = "(学習未実行 / 区分情報なし)"
        return
    max_len = max(len(arr) for arr in shape.values())
    ws.cell(row=1, column=1, value="区分").font = Font(bold=True)
    for c in range(1, max_len + 1):
        ws.cell(row=1, column=c + 1, value=f"M{c}").font = Font(bold=True)
    for r, (cat, arr) in enumerate(sorted(shape.items()), start=2):
        ws.cell(row=r, column=1, value=cat).font = Font(bold=True)
        for c, v in enumerate(arr, start=2):
            ws.cell(row=r, column=c, value=float(v))

    # SKU別の区分マッピング (参照用)
    if materials is not None and len(materials) > 0:
        map_start = len(shape) + 4
        ws.cell(row=map_start, column=1, value="=== SKU区分参照 ===").font = Font(bold=True)
        ws.cell(row=map_start + 1, column=1, value="sku_id").font = Font(bold=True)
        ws.cell(row=map_start + 1, column=2, value="区分").font = Font(bold=True)
        sorted_mat = materials.sort_values("sku_id")
        for i, row in enumerate(sorted_mat.itertuples(index=False), start=map_start + 2):
            ws.cell(row=i, column=1, value=str(row.sku_id))
            cat_val = getattr(row, "区分", None)
            ws.cell(row=i, column=2, value=str(cat_val) if pd.notna(cat_val) else "")


def _write_unmapped(wb, unmapped):
    ws = wb.create_sheet("未マッピング")
    if len(unmapped) == 0:
        ws["A1"] = "未マッピングSKUなし"
        return
    for r in dataframe_to_rows(unmapped, index=False, header=True):
        ws.append(r)


def _write_warnings(wb, warnings_list):
    ws = wb.create_sheet("警告ログ")
    ws["A1"] = "警告内容"
    ws["A1"].font = Font(bold=True)
    for i, w in enumerate(warnings_list, start=2):
        ws.cell(row=i, column=1, value=w)


def _write_dataframe_with_index(ws, df: pd.DataFrame, header_label: str):
    """index 付き DataFrame を書き出す."""
    ws.cell(row=1, column=1, value=header_label).font = Font(bold=True)
    if isinstance(df.columns, pd.MultiIndex):
        cols = [str(c) for c in df.columns.tolist()]
    else:
        cols = list(df.columns)
    for c, col_name in enumerate(cols, start=2):
        ws.cell(row=1, column=c, value=str(col_name)).font = Font(bold=True)
    if isinstance(df.index, pd.MultiIndex):
        idx_list = [" / ".join(str(x) for x in t) for t in df.index.tolist()]
    else:
        idx_list = [str(x) for x in df.index.tolist()]
    for r, idx_label in enumerate(idx_list, start=2):
        ws.cell(row=r, column=1, value=idx_label).font = Font(bold=True)
        for c, col_name in enumerate(cols, start=2):
            val = df.iat[r - 2, c - 2]
            if pd.isna(val):
                continue
            if isinstance(val, (int, float)):
                ws.cell(row=r, column=c, value=float(val))
            else:
                ws.cell(row=r, column=c, value=val)


def _write_dataframe_with_index_at(ws, df: pd.DataFrame, header_label: str, start_row: int):
    """指定行から index 付き DataFrame を書き出す."""
    ws.cell(row=start_row, column=1, value=header_label).font = Font(bold=True)
    if isinstance(df.columns, pd.MultiIndex):
        cols = [str(c) for c in df.columns.tolist()]
    else:
        cols = list(df.columns)
    for c, col_name in enumerate(cols, start=2):
        ws.cell(row=start_row, column=c, value=str(col_name)).font = Font(bold=True)
    if isinstance(df.index, pd.MultiIndex):
        idx_list = [" / ".join(str(x) for x in t) for t in df.index.tolist()]
    else:
        idx_list = [str(x) for x in df.index.tolist()]
    for r, idx_label in enumerate(idx_list, start=start_row + 1):
        ws.cell(row=r, column=1, value=idx_label).font = Font(bold=True)
        for c, col_name in enumerate(cols, start=2):
            val = df.iat[r - start_row - 1, c - 2]
            if pd.isna(val):
                continue
            if isinstance(val, (int, float)):
                ws.cell(row=r, column=c, value=float(val))
            else:
                ws.cell(row=r, column=c, value=val)
