"""_1.xlsx と類似レイアウトの予測在庫管理表 (_RES.xlsx) を生成.

月ブロック=5列: 発注済量(確定), 入庫(予測), 出庫(予測), 在庫(予測), 在庫評価(予測).
"""
import math
from datetime import date
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


_SHEET_NAME = "在庫評価（製作所、在庫していない溶材削除）"
_TITLE = "溶接棒在庫管理表 新来島どっく 大西 (予測)"
_DEFAULT_LOCATION = "大西"

_EVAL_FILL = {
    1: PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    2: PatternFill(start_color="FFD8A8", end_color="FFD8A8", fill_type="solid"),
    3: PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    4: PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
    5: PatternFill(start_color="D9C6E5", end_color="D9C6E5", fill_type="solid"),
}

_BLOCK = 5
_SUBHEADERS = ["発注済量", "入庫", "出庫", "在庫", "在庫評価"]


def _format_size_label(diameter, unit_weight) -> str:
    def _num(v):
        if v is None:
            return ""
        s = str(v).strip()
        if s == "":
            return ""
        try:
            return f"{float(s):g}" if any(ch.isdigit() for ch in s) else ""
        except ValueError:
            import re
            m = re.search(r"\d+\.?\d*", s)
            return m.group(0) if m else ""

    d = _num(diameter)
    w = _num(unit_weight)
    if d and w:
        return f"{d}× {w}"
    if d:
        return f"{d}×"
    return ""


def _isnan(v) -> bool:
    return isinstance(v, float) and math.isnan(v)


def write_inventory_format(
    out_path: str,
    orders: pd.DataFrame,
    materials: pd.DataFrame,
) -> None:
    """orders を 5列ブロックレイアウトで RES.xlsx に出力.

    Args:
        out_path: 出力ファイルパス
        orders: recommend_orders の戻り値
                (year_month, sku_id, 発注済量, 推奨発注量, 予測使用量,
                 月末予測在庫, 予測在庫評価, 手配済み, ...)
        materials: read_materials_and_consumption の戻り値の1番目
    """
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(_SHEET_NAME)

    if len(orders) == 0:
        ws["A1"] = "(予測データなし)"
        wb.save(out_path)
        return

    months = sorted(orders["year_month"].unique())
    # SKU は 推奨発注量(予測) 合計が大きい順
    sku_totals = (
        orders.groupby("sku_id")["推奨発注量"].sum()
        .sort_values(ascending=False)
    )
    skus = sku_totals.index.tolist()

    mat_map = {}
    if materials is not None and len(materials) > 0:
        for _, row in materials.iterrows():
            mat_map[row["sku_id"]] = row

    ws.cell(row=1, column=2, value=_TITLE).font = Font(bold=True, size=12)
    ws.cell(row=2, column=2, value="銘柄").font = Font(bold=True)
    ws.cell(row=2, column=4, value="納品先").font = Font(bold=True)
    for i, ym in enumerate(months):
        col = 5 + i * _BLOCK
        y, m = int(ym[:4]), int(ym[5:7])
        ws.cell(row=2, column=col, value=date(y, m, 1))
        ws.cell(row=2, column=col).number_format = "yyyy/m/d"

    for i in range(len(months)):
        base = 5 + i * _BLOCK
        for offset, label in enumerate(_SUBHEADERS):
            c = ws.cell(row=3, column=base + offset, value=label)
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center")

    orders_idx = orders.set_index(["sku_id", "year_month"])
    row_idx = 4
    for serial, sku in enumerate(skus, start=1):
        mat = mat_map.get(sku)
        if mat is not None:
            brand = mat.get("銘柄", "")
            size = _format_size_label(mat.get("径"), mat.get("単位重量"))
            ws.cell(row=row_idx, column=1, value=serial)
            ws.cell(row=row_idx, column=2, value=str(brand) if brand is not None else "")
            ws.cell(row=row_idx, column=3, value=size)
        else:
            ws.cell(row=row_idx, column=1, value=serial)
            ws.cell(row=row_idx, column=2, value=sku)
        ws.cell(row=row_idx, column=4, value=_DEFAULT_LOCATION)

        for i, ym in enumerate(months):
            base = 5 + i * _BLOCK
            if (sku, ym) in orders_idx.index:
                r = orders_idx.loc[(sku, ym)]
                confirmed = r.get("発注済量")
                if confirmed is not None and not _isnan(confirmed):
                    ws.cell(row=row_idx, column=base + 0, value=float(confirmed))
                ws.cell(row=row_idx, column=base + 1, value=float(r["推奨発注量"]))
                ws.cell(row=row_idx, column=base + 2, value=float(r["予測使用量"]))
                ws.cell(row=row_idx, column=base + 3, value=float(r["月末予測在庫"]))
                eval_val = r["予測在庫評価"]
                if pd.notna(eval_val):
                    cell = ws.cell(row=row_idx, column=base + 4, value=int(eval_val))
                    if int(eval_val) in _EVAL_FILL:
                        cell.fill = _EVAL_FILL[int(eval_val)]
        row_idx += 1

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 8
    for i in range(len(months) * _BLOCK):
        col_letter = ws.cell(row=2, column=5 + i).column_letter
        ws.column_dimensions[col_letter].width = 9

    wb.save(out_path)
