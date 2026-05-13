"""Excel入力ファイルの読み込み・正規化."""
from datetime import date
from dateutil.relativedelta import relativedelta
import pandas as pd
from openpyxl import load_workbook


def _to_date(v):
    """セル値を date に変換 (datetime/date/str いずれも対応)."""
    if v is None or v == "":
        return None
    if hasattr(v, "date"):
        return v.date()
    if isinstance(v, date):
        return v
    return pd.to_datetime(v).date()


def read_senpyou(path: str) -> pd.DataFrame:
    """線表シートを読み込み、3ドック横並びを縦持ちに変換.

    Returns: ships DataFrame
      columns: s_no, ship_type, dock, keel_date, launch_date, delivery_date,
               consume_start, consume_end
    """
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["線表"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    dock_row = rows[0]
    dock_cols = [
        (1, dock_row[0]),
        (6, dock_row[5]),
        (11, dock_row[10]),
    ]

    ships = []
    for start_col, dock_name in dock_cols:
        if dock_name is None:
            continue
        for row in rows[2:]:
            s_no = row[start_col - 1]
            if s_no is None or str(s_no).strip() == "":
                continue
            ship_type = row[start_col]
            keel = _to_date(row[start_col + 1])
            launch = _to_date(row[start_col + 2])
            delivery = _to_date(row[start_col + 3])
            if keel is None or launch is None:
                continue
            consume_start = keel - relativedelta(months=2)
            consume_end = launch + relativedelta(months=1)
            ships.append({
                "s_no": str(s_no).strip(),
                "ship_type": str(ship_type).strip() if ship_type else "",
                "dock": str(dock_name).strip(),
                "keel_date": keel,
                "launch_date": launch,
                "delivery_date": delivery,
                "consume_start": consume_start,
                "consume_end": consume_end,
            })
    return pd.DataFrame(ships)


def read_ship_types(path: str) -> pd.DataFrame:
    """船種マスタシートを読み込む.

    Returns: ship_types DataFrame
      columns: ship_type, category
    """
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["船種"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return pd.DataFrame(columns=["ship_type", "category"])

    categories = list(rows[0])
    records = []
    for row in rows[1:]:
        for col_idx, cell in enumerate(row):
            if cell is None or str(cell).strip() == "":
                continue
            if col_idx >= len(categories) or categories[col_idx] is None:
                continue
            records.append({
                "ship_type": str(cell).strip(),
                "category": str(categories[col_idx]).strip(),
            })
    return pd.DataFrame(records)


from src.sku_normalizer import make_sku_id


_AGGREGATE_BRANDS = {"CO2", "SEG", "裏当材", "溶接棒"}


def detect_train_end_ym(inventory: pd.DataFrame,
                        consumption: pd.DataFrame) -> str | None:
    """学習期間の終端 year_month を自動判定.

    判定基準:
      - inventory_latest: 在庫管理表で「出庫 または 在庫 が非ゼロ」のレコードが存在する最新月
      - consumption_latest: 消費データで「qty_kg が非ゼロ」のレコードが存在する最新月
      - train_end = min(inventory_latest, consumption_latest)

    どちらか片方が空なら、 もう片方の最新月を採用. 両方空なら None.
    """
    inv_latest = None
    if len(inventory) > 0:
        valid = inventory[(inventory["出庫"].fillna(0) > 0)
                          | (inventory["在庫"].fillna(0) > 0)]
        if len(valid) > 0:
            inv_latest = valid["year_month"].max()
    cons_latest = None
    if len(consumption) > 0:
        valid = consumption[consumption["qty_kg"].fillna(0) > 0]
        if len(valid) > 0:
            cons_latest = valid["year_month"].max()

    if inv_latest is None and cons_latest is None:
        return None
    if inv_latest is None:
        return cons_latest
    if cons_latest is None:
        return inv_latest
    return min(inv_latest, cons_latest)


def _is_aggregate_column(brand: str, category) -> bool:
    """集計・合計・単位換算列を判別.

    以下の場合は集計/合計/単位換算列として除外する:
    - 銘柄が集計カテゴリ名そのもの (CO2, SEG, 裏当材, 溶接棒)
    - 単位換算行 ("裏当て1本300グラム" など)
    - 区分が空 (実SKUは A-I のいずれか)
    """
    # 銘柄が集計カテゴリ名そのもの
    if brand in _AGGREGATE_BRANDS:
        return True
    # 単位換算行 ("裏当て1本300グラム" 等)
    if "裏当て" in brand and "グラム" in brand:
        return True
    # 区分が空 (実SKUは A-I のいずれか)
    cat_str = str(category).strip() if category is not None else ""
    if cat_str == "":
        return True
    return False


def read_materials_and_consumption(path: str):
    """溶接材料シートを読み込み、materials と monthly_consumption を返す.

    Returns: (materials_df, consumption_df)
    """
    wb = load_workbook(path, data_only=True, read_only=False)
    ws = wb["溶接材料"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 6:
        return pd.DataFrame(), pd.DataFrame()

    row1, row2, row3, row4, row5 = rows[0], rows[1], rows[2], rows[3], rows[4]
    materials = []
    sku_id_by_col = {}
    n_cols = max(len(row1), len(row2), len(row3), len(row4))
    for c in range(2, n_cols):
        brand = row2[c] if c < len(row2) else None
        diameter = row3[c] if c < len(row3) else None
        weight = row4[c] if c < len(row4) else None
        category = row5[c] if c < len(row5) else None
        if brand is None or str(brand).strip() == "":
            continue

        # FILTER: skip subtotal / aggregate / unit-conversion columns
        brand_str = str(brand).strip()
        if _is_aggregate_column(brand_str, category):
            continue

        sku_id = make_sku_id(brand, diameter, weight)
        sku_id_by_col[c] = sku_id
        materials.append({
            "sku_id": sku_id,
            "銘柄": str(brand).strip(),
            "径": diameter,
            "単位重量": weight,
            "区分": row5[c] if c < len(row5) else "",
            "溶接法": row1[c] if c < len(row1) else "",
        })

    consumption = []
    for r in range(5, len(rows)):
        ym_raw = rows[r][1] if len(rows[r]) > 1 else None
        if ym_raw is None or str(ym_raw).strip() == "":
            continue
        ym_str = _normalize_year_month(ym_raw)
        if ym_str is None:
            continue
        for c, sku_id in sku_id_by_col.items():
            if c >= len(rows[r]):
                continue
            val = rows[r][c]
            qty = _parse_quantity(val)
            consumption.append({
                "year_month": ym_str,
                "sku_id": sku_id,
                "qty_kg": qty,
            })

    return pd.DataFrame(materials), pd.DataFrame(consumption)


def _normalize_year_month(raw) -> str:
    """'2017/01' or datetime → 'YYYY-MM' に統一."""
    if hasattr(raw, "year"):
        return f"{raw.year:04d}-{raw.month:02d}"
    s = str(raw).strip()
    parts = s.replace("-", "/").split("/")
    if len(parts) >= 2:
        try:
            return f"{int(parts[0]):04d}-{int(parts[1]):02d}"
        except ValueError:
            return None
    return None


def _parse_quantity(v) -> float:
    """セル値を数値に: None/空 → 0、カンマ付き文字列 → float."""
    if v is None or str(v).strip() == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return 0.0


import re
from src.sku_normalizer import parse_inventory_name


_INVENTORY_SHEET_PATTERN = re.compile(r"在庫評価")


def _detect_month_columns(row2):
    """ヘッダ第2行から月列の開始位置とラベルを抽出."""
    month_cols = {}
    for c, val in enumerate(row2):
        if val is None:
            continue
        if hasattr(val, "year") and hasattr(val, "month"):
            month_cols[c] = f"{val.year:04d}-{val.month:02d}"
        else:
            m = re.match(r"(\d{4})年\s*(\d{1,2})月", str(val))
            if m:
                month_cols[c] = f"{int(m.group(1)):04d}-{int(m.group(2)):02d}"
    return month_cols


def _detect_block_width(row3, month_cols):
    """月ブロックの幅（列数）を推定。

    _1.xlsx 系: (入庫, 出庫, 在庫, 在庫評価) → 4列
    _2.xlsx 系: (入荷予定, 出庫, 在庫) → 3列
    """
    cols = sorted(month_cols.keys())
    if len(cols) >= 2:
        return cols[1] - cols[0]
    if len(cols) == 1:
        # 第3行で当該月ブロックの右側にラベルがいくつ並んでいるかを数える
        start = cols[0]
        end = start
        for c in range(start, len(row3)):
            v = row3[c]
            if v is None or str(v).strip() == "":
                break
            end = c
        return max(3, end - start + 1)
    return 4


def read_planned_orders(path: str) -> pd.DataFrame:
    """検証用ファイル (_2.xlsx) の入荷予定をSKU×月で抽出.

    _2.xlsx は3列構造 (入荷予定/出庫/在庫) で月ブロックが並ぶ。
    出庫列にはテキストメモが入る場合があるので数値以外は無視。
    入荷予定 (各月ブロックの先頭列) のみを数値として読み取る。

    Returns: DataFrame columns = year_month (str), sku_id (str), planned_kg (float)
    """
    wb = load_workbook(path, data_only=True, read_only=True)
    target_sheet = None
    for name in wb.sheetnames:
        if _INVENTORY_SHEET_PATTERN.search(name):
            target_sheet = name
            break
    if target_sheet is None:
        wb.close()
        return pd.DataFrame(columns=["year_month", "sku_id", "planned_kg"])
    ws = wb[target_sheet]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 4:
        return pd.DataFrame(columns=["year_month", "sku_id", "planned_kg"])

    month_cols = _detect_month_columns(rows[1])
    if not month_cols:
        return pd.DataFrame(columns=["year_month", "sku_id", "planned_kg"])

    records = []
    last_sku_id = None
    for r in range(3, len(rows)):
        row = rows[r]
        brand_part1 = row[1] if len(row) > 1 else None
        brand_part2 = row[2] if len(row) > 2 else None
        location = row[3] if len(row) > 3 else None
        if (brand_part1 is None or str(brand_part1).strip() == "") and \
           (brand_part2 is None or str(brand_part2).strip() == ""):
            if location is None or str(location).strip() == "":
                continue
            if last_sku_id is None:
                continue
            sku_id = last_sku_id
        else:
            combined = f"{brand_part1 or ''} {brand_part2 or ''}".strip()
            # 集計行 (B列が「合計」「総合計」「総計」など) を除外
            if any(kw in combined for kw in ("合計", "総計", "計")) and \
               (brand_part2 is None or str(brand_part2).strip() == ""):
                last_sku_id = None
                continue
            sku_id, _brand = parse_inventory_name(combined)
            last_sku_id = sku_id
        # sku_id が "_" のみで構成される無効値 (銘柄部分が空) は除外
        if sku_id is None or sku_id.replace("_", "") == "":
            continue
        for col_idx, ym in month_cols.items():
            v = row[col_idx] if col_idx < len(row) else None
            qty = _parse_quantity(v)
            if qty == 0:
                continue
            records.append({
                "year_month": ym,
                "sku_id": sku_id,
                "planned_kg": qty,
            })
    if not records:
        return pd.DataFrame(columns=["year_month", "sku_id", "planned_kg"])
    df = pd.DataFrame(records)
    return df.groupby(["year_month", "sku_id"], as_index=False)["planned_kg"].sum()


def read_inventory(path: str) -> pd.DataFrame:
    """在庫管理表ファイルの「在庫評価」シートを読み込む.

    Returns: inventory DataFrame
      columns: year_month, sku_id, 入庫, 出庫, 在庫, 拠点, 在庫評価

    注意: 銘柄が空で 拠点(D列) のみ入っている行は「直前銘柄の別拠点行」
    """
    wb = load_workbook(path, data_only=True, read_only=True)
    target_sheet = None
    for name in wb.sheetnames:
        if _INVENTORY_SHEET_PATTERN.search(name):
            target_sheet = name
            break
    if target_sheet is None:
        wb.close()
        return pd.DataFrame()
    ws = wb[target_sheet]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 4:
        return pd.DataFrame()

    month_cols = {}
    row2 = rows[1]
    for c, val in enumerate(row2):
        if val is None:
            continue
        # Handle both datetime and string formats
        if hasattr(val, "year") and hasattr(val, "month"):
            # datetime object
            year = val.year
            month = val.month
            month_cols[c] = f"{year:04d}-{month:02d}"
        else:
            # String format like "2024年8月"
            m = re.match(r"(\d{4})年\s*(\d{1,2})月", str(val))
            if m:
                year = int(m.group(1))
                month = int(m.group(2))
                month_cols[c] = f"{year:04d}-{month:02d}"

    records = []
    last_sku_id = None
    for r in range(3, len(rows)):
        row = rows[r]
        brand_part1 = row[1] if len(row) > 1 else None
        brand_part2 = row[2] if len(row) > 2 else None
        location = row[3] if len(row) > 3 else None
        # B/C 列が空 && D 列に拠点がある → 直前銘柄を継承
        if (brand_part1 is None or str(brand_part1).strip() == "") and \
           (brand_part2 is None or str(brand_part2).strip() == ""):
            if location is None or str(location).strip() == "":
                continue
            if last_sku_id is None:
                continue
            sku_id = last_sku_id
        else:
            combined = f"{brand_part1 or ''} {brand_part2 or ''}".strip()
            sku_id, _brand = parse_inventory_name(combined)
            last_sku_id = sku_id
        loc = str(location).strip() if location else ""
        for col_idx, ym in month_cols.items():
            ny = row[col_idx] if col_idx < len(row) else None
            de = row[col_idx + 1] if col_idx + 1 < len(row) else None
            za = row[col_idx + 2] if col_idx + 2 < len(row) else None
            ev = row[col_idx + 3] if col_idx + 3 < len(row) else None
            if ny is None and de is None and za is None and ev is None:
                continue
            records.append({
                "year_month": ym,
                "sku_id": sku_id,
                "入庫": _parse_quantity(ny),
                "出庫": _parse_quantity(de),
                "在庫": _parse_quantity(za),
                "在庫評価": _parse_quantity(ev) if ev is not None else None,
                "拠点": loc,
            })
    return pd.DataFrame(records)


def extract_planned_orders_from_inventory(inventory: pd.DataFrame) -> pd.DataFrame:
    """在庫管理表 (_1.xlsx) の「手配済み入庫のみ」行を抽出.

    定義: 入庫 > 0 かつ 出庫 = 0 かつ 在庫 = 0
    (実績の出庫/在庫が未登録だが、入庫予定だけが計上された行)

    銘柄部が空 (sku_id が `_<diameter>_<weight>` パターン) の行は除外。
    例: 小計行や空セル由来の sku_id="__"。

    Returns: DataFrame
      columns: year_month (str), sku_id (str), planned_kg (float)
    """
    empty = pd.DataFrame(columns=["year_month", "sku_id", "planned_kg"])
    if len(inventory) == 0:
        return empty
    mask = (
        (inventory["入庫"].fillna(0) > 0)
        & (inventory["出庫"].fillna(0) == 0)
        & (inventory["在庫"].fillna(0) == 0)
    )
    # sku_id の銘柄部 (`_` 区切り先頭) が空の行を除外
    brand_part = inventory["sku_id"].astype(str).str.split("_", n=1).str[0]
    mask = mask & (brand_part != "")
    if not mask.any():
        return empty
    out = inventory.loc[mask, ["year_month", "sku_id", "入庫"]].copy()
    out = out.rename(columns={"入庫": "planned_kg"})
    out = (
        out.groupby(["year_month", "sku_id"], as_index=False)["planned_kg"].sum()
    )
    return out


def read_forecast_result(path: str) -> pd.DataFrame:
    """RES.xlsx (5列ブロック: 発注済量/入庫/出庫/在庫/在庫評価) を読み込む.

    Returns: DataFrame columns:
      year_month, sku_id, 発注済量, 入庫_予測, 出庫_予測, 在庫_予測, 在庫評価, 拠点

    注意:
      - 発注済量はブランクなら NaN
      - 銘柄が空で 拠点(D列) のみある行は直前銘柄の別拠点
    """
    wb = load_workbook(path, data_only=True, read_only=True)
    target_sheet = None
    for name in wb.sheetnames:
        if _INVENTORY_SHEET_PATTERN.search(name):
            target_sheet = name
            break
    if target_sheet is None:
        wb.close()
        return pd.DataFrame()
    ws = wb[target_sheet]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 4:
        return pd.DataFrame()

    month_cols = _detect_month_columns(rows[1])
    if not month_cols:
        return pd.DataFrame()

    records = []
    last_sku_id = None
    for r in range(3, len(rows)):
        row = rows[r]
        brand_part1 = row[1] if len(row) > 1 else None
        brand_part2 = row[2] if len(row) > 2 else None
        location = row[3] if len(row) > 3 else None
        if (brand_part1 is None or str(brand_part1).strip() == "") and \
           (brand_part2 is None or str(brand_part2).strip() == ""):
            if location is None or str(location).strip() == "":
                continue
            if last_sku_id is None:
                continue
            sku_id = last_sku_id
        else:
            combined = f"{brand_part1 or ''} {brand_part2 or ''}".strip()
            sku_id, _brand = parse_inventory_name(combined)
            last_sku_id = sku_id
        loc = str(location).strip() if location else ""
        for col_idx, ym in month_cols.items():
            placed = row[col_idx] if col_idx < len(row) else None
            nyu = row[col_idx + 1] if col_idx + 1 < len(row) else None
            de = row[col_idx + 2] if col_idx + 2 < len(row) else None
            za = row[col_idx + 3] if col_idx + 3 < len(row) else None
            ev = row[col_idx + 4] if col_idx + 4 < len(row) else None
            if (placed is None and nyu is None and de is None
                    and za is None and ev is None):
                continue
            records.append({
                "year_month": ym,
                "sku_id": sku_id,
                "発注済量": _parse_quantity(placed) if placed is not None else float("nan"),
                "入庫_予測": _parse_quantity(nyu),
                "出庫_予測": _parse_quantity(de),
                "在庫_予測": _parse_quantity(za),
                "在庫評価": _parse_quantity(ev) if ev is not None else None,
                "拠点": loc,
            })
    return pd.DataFrame(records)
